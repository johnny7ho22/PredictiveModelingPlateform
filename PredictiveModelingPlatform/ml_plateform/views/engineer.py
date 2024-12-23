from django.shortcuts import render,redirect,HttpResponse
from openpyxl import load_workbook
from ml_plateform import models
from ml_plateform.utils.BootstrapModelForm import BootstrapModelForm 
from ml_plateform.utils.pagination import Pagination
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
import io
import base64
# macOS 不允許Matplotlib 嘗試在非主線程中啟動圖形界面 (GUI)
# 如果不需要顯示圖形窗口（GUI），可以切換到一個非交互式的後端（如 Agg），這樣就不會啟動 GUI 操作,這樣才不會報錯
import matplotlib
matplotlib.use('Agg')  
from django.http import JsonResponse
import numpy as np
from scipy.stats import gaussian_kde
from django.views.decorators.csrf import csrf_exempt 
from sklearn.preprocessing import StandardScaler,LabelEncoder,MinMaxScaler,PowerTransformer
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import accuracy_score, confusion_matrix, classification_report
import warnings
warnings.filterwarnings("ignore")





class DataModelForm(BootstrapModelForm):
    class Meta:
        model = models.HeartFailure
        fields = "__all__"





def upload_data(request):
    if request.method == "GET":
        return render(request, "engineer_upload.html")
    
    file_object = request.FILES.get("exc")

    wb = load_workbook(file_object)
    sheet = wb.worksheets[0]

    for row in sheet.iter_rows(min_row=2):
        Age = row[0].value
        Sex = row[1].value
        ChestPainType = row[2].value
        FastingBS = row[3].value
        MaxHR = row[4].value
        ExerciseAngina = row[5].value
        Oldpeak = row[6].value
        ST_Slope = row[7].value
        HeartDisease = row[8].value

        models.HeartFailure.objects.create(Age = Age, Sex = Sex, ChestPainType = ChestPainType, 
                                           FastingBS = FastingBS, MaxHR = MaxHR, ExerciseAngina = ExerciseAngina,
                                           Oldpeak = Oldpeak, ST_Slope = ST_Slope, HeartDisease = HeartDisease)


    return redirect("/engineer/list")






def visualization(request):
    """資料視覺化"""

    return render(request, 'engineer_visualization.html')






def show_corr(df):
    plt.figure(figsize=(15,5),facecolor='tan')
    sns.heatmap(df.corr(),annot=True, cmap="inferno")
    plt.title("Correlation Plot of the Heart Failure Prediction")

    buffer = io.BytesIO()
    plt.savefig(buffer, format = "png")
    buffer.seek(0)
    image_png = buffer.getvalue()
    buffer.close()
    corr_img = base64.b64encode(image_png).decode("utf-8")

    return corr_img



def chart_imbalance(request):
    """資料集是否平衡"""


    dataset = models.HeartFailure.objects.all()

    dataset_list = list(dataset.values())

    df = pd.DataFrame(dataset_list)

    total = df.shape[0]

    target_counts = df['HeartDisease'].value_counts() #Serires

    target_counts_df = pd.DataFrame(target_counts).reset_index()

    target_counts_df.columns = ['HeartDisease', 'Count']

    data = [{"name": "Yes", "y": target_counts_df.iloc[0,1]/total*100}, {"name": "No", "y": target_counts_df.iloc[1,1]/total*100}]

    result = {"status":True, "data":data}



    return JsonResponse(result)



def sex_disease(request):
    """查看男女性患有心臟病之分佈"""
    male_normal_count = models.HeartFailure.objects.filter(Sex = 1, HeartDisease = 0).count()
    male_disease_count = models.HeartFailure.objects.filter(Sex = 1, HeartDisease = 1).count()
    female_normal_count = models.HeartFailure.objects.filter(Sex = 0, HeartDisease = 0).count()
    female_disease_count = models.HeartFailure.objects.filter(Sex = 0, HeartDisease = 1).count()

    categories = ['無','有']

    series = [
        {"name" : "男性", "data":[male_normal_count,male_disease_count]},
        {"name" : "女性", "data":[female_normal_count,female_disease_count]}

    ]
    

    result = {"status":True, "categories":categories, "series":series, "title":"男女性患有心臟病之分佈"}

    return JsonResponse(result)



def chest_pain(request):
    """男、女性胸口疼痛類型分佈"""

    # 0 : ASY, 1 : ATA, 2 : NAP, 3 :TA
    male_ASY_count = models.HeartFailure.objects.filter(Sex = 1, ChestPainType = 0).count()
    male_ATA_count = models.HeartFailure.objects.filter(Sex = 1, ChestPainType = 1).count()
    male_NAP_count = models.HeartFailure.objects.filter(Sex = 1, ChestPainType = 2).count()
    male_TA_count = models.HeartFailure.objects.filter(Sex = 1, ChestPainType = 3).count()

    female_ASY_count = models.HeartFailure.objects.filter(Sex = 0, ChestPainType = 0).count()
    female_ATA_count = models.HeartFailure.objects.filter(Sex = 0, ChestPainType = 1).count()
    female_NAP_count = models.HeartFailure.objects.filter(Sex = 0, ChestPainType = 2).count()
    female_TA_count = models.HeartFailure.objects.filter(Sex = 0, ChestPainType = 3).count()

    categories = ['ASY','ATA','NAP','TA']

    series = [
        {"name" : "男性", "data":[male_ASY_count,male_ATA_count,male_NAP_count,male_TA_count]},
        {"name" : "女性", "data":[female_ASY_count,female_ATA_count,female_NAP_count,female_TA_count]}
    ]
    

    result = {"status":True, "categories":categories, "series":series, "title":"男、女性胸口疼痛類型分佈"}

    return JsonResponse(result)










def box_plot(request):

    age = list(models.HeartFailure.objects.values_list('Age', flat=True))
    maxHr = list(models.HeartFailure.objects.values_list('MaxHR', flat=True))
    oldPeak = list(models.HeartFailure.objects.values_list('Oldpeak', flat=True))

    # 計算四分位數
    age_Q1 = np.percentile(age, 25)  # 第一四分位數
    age_Q2 = np.percentile(age, 50)  # 第二四分位數 (中位數)
    age_Q3 = np.percentile(age, 75)  # 第三四分位數
    age_IQR = age_Q3 - age_Q1
    age_minimum = age_Q1 - 1.5 * age_IQR
    age_maximum = age_Q3 + 1.5 * age_IQR

    age_data = [age_minimum, age_Q1, age_Q2, age_Q3, age_maximum]


    maxHr_Q1 = np.percentile(maxHr, 25)  # 第一四分位數
    maxHr_Q2 = np.percentile(maxHr, 50)  # 第二四分位數 (中位數)
    maxHr_Q3 = np.percentile(maxHr, 75)  # 第三四分位數
    maxHr_IQR = maxHr_Q3 - maxHr_Q1
    maxHr_minimum = maxHr_Q1 - 1.5 * maxHr_IQR
    maxHr_maximum = maxHr_Q3 + 1.5 * maxHr_IQR

    maxHr_data = [maxHr_minimum, maxHr_Q1, maxHr_Q2, maxHr_Q3, maxHr_maximum]


    oldPeak_Q1 = np.percentile(oldPeak, 25)  # 第一四分位數
    oldPeak_Q2 = np.percentile(oldPeak, 50)  # 第二四分位數 (中位數)
    oldPeak_Q3 = np.percentile(oldPeak, 75)  # 第三四分位數
    oldPeak_IQR = oldPeak_Q3 - oldPeak_Q1
    oldPeak_minimum = oldPeak_Q1 - 1.5 * oldPeak_IQR
    oldPeak_maximum = oldPeak_Q3 + 1.5 * oldPeak_IQR

    oldPeak_data = [oldPeak_minimum, oldPeak_Q1, oldPeak_Q2, oldPeak_Q3, oldPeak_maximum]


    data = [age_data, maxHr_data, oldPeak_data]

    result = {
        "status":True,
        "feature":["Age", "MaxHR", "OldPeak"],
        "data":data,
    }

    return JsonResponse(result)


def cfeature_target(request):
    """類別型特徵對目標變數的影響"""

    data_list = []
    titles = ["男、女性換患有心臟病之比例", 
              "胸口疼痛型態對患有心臟病之比例", 
              "FastingBS對患有心臟病之比例", 
              "ExerciseAngina對患有心臟病之比例",
              "ST_Slope對患有心臟病之比例",
    ]

    #性別對心臟病之影響
    sex_list = []

    male_heartDisease = models.HeartFailure.objects.filter(Sex = 1, HeartDisease = 1).count()
    female_heartDisease = models.HeartFailure.objects.filter(Sex = 0, HeartDisease = 1).count()

    sex_list.append({"name":"男性", "y": male_heartDisease/ (male_heartDisease + female_heartDisease) *100})
    sex_list.append({"name":"女性", "y": female_heartDisease/ (male_heartDisease + female_heartDisease) *100})

    #胸口疼痛對心臟病之影響
    chestPainType_list = []
    
    ASY_count = models.HeartFailure.objects.filter(ChestPainType = 0, HeartDisease = 1).count()
    ATA_count = models.HeartFailure.objects.filter(ChestPainType = 1, HeartDisease = 1).count()
    NAP_count = models.HeartFailure.objects.filter(ChestPainType = 2, HeartDisease = 1).count()
    TA_count = models.HeartFailure.objects.filter(ChestPainType = 3, HeartDisease = 1).count()

    chestPainTypeTotal = ASY_count + ATA_count + NAP_count + TA_count

    chestPainType_list.append({"name":"ASY", "y": ASY_count/ chestPainTypeTotal *100})
    chestPainType_list.append({"name":"ATA", "y": ATA_count/ chestPainTypeTotal *100})
    chestPainType_list.append({"name":"NAP", "y": NAP_count/ chestPainTypeTotal *100})
    chestPainType_list.append({"name":"TA", "y": TA_count/ chestPainTypeTotal *100})


    #FastingBS對心臟病之影響
    FastingBS_list = []
    
    over_count = models.HeartFailure.objects.filter(FastingBS = 0, HeartDisease = 1).count()
    under_count = models.HeartFailure.objects.filter(FastingBS = 1, HeartDisease = 1).count()

    FastingBSTotal = over_count + under_count 


    FastingBS_list.append({"name":"FBS < 120", "y": over_count/ FastingBSTotal *100})
    FastingBS_list.append({"name":"FBS > 120", "y": under_count/ FastingBSTotal *100})


    #ExerciseAngina對心臟病之影響
    ExerciseAngina_list = []
    
    no_Angina_count = models.HeartFailure.objects.filter(ExerciseAngina = 0, HeartDisease = 1).count()
    Angina_count = models.HeartFailure.objects.filter(ExerciseAngina = 1, HeartDisease = 1).count()

    ExerciseAnginaTotal = Angina_count + no_Angina_count

    ExerciseAngina_list.append({"name":"No Angina", "y": no_Angina_count/ ExerciseAnginaTotal *100})
    ExerciseAngina_list.append({"name":"Angina", "y": Angina_count/ ExerciseAnginaTotal *100})


    #ST_Slope對心臟病之影響
    ST_Slope_list = []

    Down_count = models.HeartFailure.objects.filter(ST_Slope = 0, HeartDisease = 1).count()
    Flat_count = models.HeartFailure.objects.filter(ST_Slope = 1, HeartDisease = 1).count()
    Up_count = models.HeartFailure.objects.filter(ST_Slope = 2, HeartDisease = 1).count()

    ST_SlopeTotal = Flat_count + Down_count + Up_count 

    ST_Slope_list.append({"name":"Flat", "y": Flat_count/ ST_SlopeTotal *100})
    ST_Slope_list.append({"name":"Down", "y": Down_count/ ST_SlopeTotal *100})
    ST_Slope_list.append({"name":"up", "y": Up_count/ ST_SlopeTotal *100})





    data_list = [sex_list, chestPainType_list, FastingBS_list, ExerciseAngina_list,ST_Slope_list]

    result = {
        "status":True,
        "titles":titles,
        "data_list":data_list
    }
    

    return JsonResponse(result)


def model_page(request):
    """產生模型頁面與資料集"""

    form = DataModelForm()

    queryset = models.HeartFailure.objects.all()

    page_object = Pagination(request,queryset)

    #分頁後的數據
    page_queryset = page_object.page_queryset

    #生成頁碼
    page_string = page_object.html()


    context =  {"form" : form, "queryset": page_queryset, "page_string": page_string}

    
    return render(request, "engineer_page.html", context)



@csrf_exempt
def model_add(request):
    """新增資料"""

    form = DataModelForm(data = request.POST)

    if form.is_valid():
        form.save()

        return JsonResponse({"status": True})
    

    return JsonResponse({"status":False, 'error':form.errors})


def model_delete(requset):
    """刪除資料"""
    uid = requset.GET.get("uid")
    exists = models.HeartFailure.objects.filter(id = uid).exists()

    if not exists:
        return JsonResponse({"status":False, "error": "刪除失敗：數據不存在"})
    
    models.HeartFailure.objects.filter(id = uid).delete()
    
    return JsonResponse({"status":True})


def edit_detail(request):
    """取得要編輯的數據"""
    uid = request.GET.get("uid")

    row_dict = models.HeartFailure.objects.filter(id = uid).values("Age", "Sex", "ChestPainType","FastingBS","MaxHR","ExerciseAngina","Oldpeak","ST_Slope","HeartDisease").first()

    if not row_dict:
        return JsonResponse({"status":False, "error": "該數據已經不存在"})
    
    result = {
        "status":True,
        "data": row_dict
    }

    
    return JsonResponse(result)


@csrf_exempt
def model_edit(request):
    """編輯數據"""

    uid = request.GET.get("uid")
    
    row_object = models.HeartFailure.objects.filter(id = uid).first()

    if not row_object:      
        return JsonResponse({"status":False, "tips": "該數據已經不存在"})
    
    form = DataModelForm(data=request.POST, instance = row_object)

    if form.is_valid():
        form.save()
        return JsonResponse({"status":True})
    
    return JsonResponse({"status":False, "error": form.errors})



class PredictModelForm(BootstrapModelForm):
    class Meta:
        model = models.HeartFailure
        exclude = ['HeartDisease']
    



@csrf_exempt
def model_predict(request):
    """模型預測"""


    #建立模型
    dataset = models.HeartFailure.objects.all()

    dataset_list = list(dataset.values())

    df = pd.DataFrame(dataset_list)

    numerical_features = ["Age","MaxHR","Oldpeak"]

    scalar = StandardScaler()

    for feature in numerical_features:
        df[feature] = scalar.fit_transform(df[feature].values.reshape(-1,1))

    X = df.drop(["id","HeartDisease"], axis = 1)
    y = pd.DataFrame(df["HeartDisease"])

    print("欄位:",X.columns)

    X_train, X_test, y_train, y_test = train_test_split(X, y, test_size=0.2, random_state=42)


    model = LogisticRegression(random_state=0,C=10,penalty = "l2")

    model.fit(X_train, y_train)

    y_pred = model.predict(X_test)

    accuracy = accuracy_score(y_test, y_pred)

    print("模型正確率；", accuracy)


    #預測用戶輸入資料
    form = PredictModelForm(data = request.POST)

    if form.is_valid():
        cleaned_data = form.cleaned_data
        Age = cleaned_data.get("Age")
        Sex = cleaned_data.get("Sex")
        ChestPainType = cleaned_data.get("ChestPainType")
        FastingBS = cleaned_data.get("FastingBS")
        MaxHR = cleaned_data.get("MaxHR")
        ExerciseAngina = cleaned_data.get("ExerciseAngina")
        Oldpeak = cleaned_data.get("Oldpeak")
        ST_Slope = cleaned_data.get("ST_Slope")

        new_data = [[Age, Sex, ChestPainType, FastingBS, MaxHR, ExerciseAngina, Oldpeak, ST_Slope]]
        print("新資料: ",new_data)


        user_input_prediction = model.predict(new_data).tolist()

        print("現在用戶預測：", user_input_prediction[0])


        context = {
            "status":True,
            "accuracy": round(accuracy * 100,2) ,
            "user_input_prediction":user_input_prediction           
        }

        return JsonResponse(context)

    return JsonResponse({"status":False, "error": form.errors})





